'''
@author king_xiong
@date 2022-03-10 11:58
'''
from time import sleep

from openpyxl import Workbook
import json

import requests
from selenium import webdriver


def get_content(url, headers, pages):
    '''
    # 获取页面内容
    :return:
    '''
    content = requests.get(url=url, headers=headers, timeout=10).text
    # print(content)
    json_content = json.loads(content)
    # print(json_content)
    # 获取hotelid
    # hotelid_content = json_content["data"]["hotelList"]
    for i in range(20):
        hotelId = json_content["data"]["hotelList"][i]["hotelId"]
        hotelName = json_content["data"]["hotelList"][i]["hotelName"]
        starLevelDes = json_content["data"]["hotelList"][i]["starLevelDes"]
        commentScore = json_content["data"]["hotelList"][i]["commentScore"]
        commentScoreDes = json_content["data"]["hotelList"][i]["commentScoreDes"]
        commentMainTag = json_content["data"]["hotelList"][i]["commentMainTag"]
        print(hotelName, hotelId, starLevelDes, commentScore, commentScoreDes, commentMainTag)
        ws['A' + str(pages * 20 + i + 1)].value = hotelName
        ws['B' + str(pages * 20 + i + 1)].value = starLevelDes
        ws['C' + str(pages * 20 + i + 1)].value = commentScore
        ws['D' + str(pages * 20 + i + 1)].value = commentScoreDes
        ws['E' + str(pages * 20 + i + 1)].value = commentMainTag
        ws['F' + str(pages * 20 + i + 1)].value = hotelId
        wbk.save('艺龙_长沙.xlsx')
        if ((pages * 20 + i + 1) != 0 and(pages * 20 + i + 1)%500==0):
            sleep(60)


if __name__ == '__main__':
    # 创建Excel文件
    wbk = Workbook()
    ws = wbk.active
    ws.title = '艺龙_长沙'
    ws['A1'] = 'hotelName'
    ws['B1'] = 'starLevelDes'
    ws['C1'] = 'commentScore'
    ws['D1'] = 'commentScoreDes'
    ws['E1'] = 'commentMainTag'
    ws['F1'] = 'hotelId'
    for pages in range(335):
        url = 'https://hotel.elong.com/tapi/v2/list?pageSize=20&city=1901&filterList=8888_1&pageIndex=' + str(pages)
        # 首页地址 https://hotels.ctrip.com/hotels/list?countryId=1&city=206&highPrice=-1&barCurr=CNY&location=483&sort=1&category=eyJmaWx0ZXJJZCI6Ijc1fFRBR180ODkiLCJ0eXBlIjoiNzUiLCJ2YWx1ZSI6IjQ4OSIsInN1YlR5cGUiOiIyIiwiY2hpbGRWYWx1ZSI6IiIsInByb3BlcnR5VmFsdWUiOiIifQ%3D%3D
        headers = {
            'Cookie': 'CookieGuid=e01b0b39-1b57-4653-9fef-7e9bcf4d023c; Esid=617bd5b2-42aa-4b1e-8851-87cf6c347152; s_visit=1; __tctma=20377580.1646888607773632.1646888607382.1646888607382.1646888607382.1; H5CookieId=fd855841-3eac-4d2f-a681-3540367a1509; SessionToken=a5e5d865-8ee0-4f33-b6c9-f5e581d79ac401; Lgid=LRpRtrsC3gsExwGXhEk%2FlpaR3waA7McUH7SGryL5%2FSFcFGoo0crBiFn5GTyj0g20q2vdiTu0hwU7azxUcdT5sKqFmkqhoV5IOnF1wz8O1t6DRCx0FNfxth2nnFvd6nP9hxu8%2F5jmrKlD83F7fYffHQ%3D%3D; tcUser=%7B%22AccessToken%22%3A%22C8DADB2BD29F0020DDFF7FBB958B439B%22%2C%22MemberId%22%3A%22177c6593036efd6bc2fadd2a2304dfa9%22%7D; __tctmb=20377580.4205298667434326.1646888607382.1646888655382.2; ShHotel=CityName=%E9%95%BF%E6%B2%99%E5%B8%82&CityNameCN=%E9%95%BF%E6%B2%99%E5%B8%82&CityNameEN=%E9%95%BF%E6%B2%99%E5%B8%82&CityID=1901&OutDate=2022-03-11&InDate=2022-03-10; CitySearchHistory=1901%23%E9%95%BF%E6%B2%99%E5%B8%82%23%E9%95%BF%E6%B2%99%E5%B8%82%23; businessLine=hotel; orderFrom=1003; firsttime=1646888668035; indate=2022-03-10; outdate=2022-03-11; ext_param=bns%3D4%26ct%3D3; H5Channel=mnoreferseo%2CSEO; lasttime=1646889036948; JSESSIONID=D06D3F20B62E4C6EFF08D53DC0FE2DC1',
            'Connection': 'keep-alive',
            'Host': 'hotel.elong.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36',
        }
        get_content(url, headers, pages)
