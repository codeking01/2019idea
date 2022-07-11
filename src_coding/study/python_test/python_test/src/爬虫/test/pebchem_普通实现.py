import json
import requests
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl


# # 引入phantomjs无界面浏览器下载最后的图片
# from selenium import webdriver
# path='phantomjs.exe'
# browser=webdriver.PhantomJS(path)

# 封装这个浏览器对象
def share_browser():
    chrome_options = Options()
    # 如果该目录不存在则直接创建 下载的东西存在 D:\A_test_pic
    prefs = {"download.default_directory": "D:\A_test_pic"}
    # 将自定义设置添加到chrome配置对象实例中
    chrome_options.add_experimental_option("prefs", prefs)

    chrome_options.add_argument('‐‐headless')
    chrome_options.add_argument('‐‐disable‐gpu')
    path = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
    chrome_options.binary_location = path
    browser = webdriver.Chrome(chrome_options=chrome_options)
    return browser


def read_excel_file(r):
    # r代表行数

    # 获取excle文件里面的内容
    wb = openpyxl.load_workbook('query.xlsx')
    # print(type(wb))
    sheet = wb['Sheet1']
    # sheet.cell(row=1,column=1).value
    # cell=sheet['A1']
    # **采集到要查询的数据 是str类型
    cell = sheet.cell(row=r, column=1).value
    # print(cell)
    if cell != None:
        # 方便后期维护
        new_cell = cell
        # query_list = [new_cell]
        r+=1
        return new_cell
    else:
        return None

def find_excel_rows():
    wb=openpyxl.load_workbook('query.xlsx')
    sheet=wb['Sheet1']
    # 获取到行数
    rows = sheet.max_row
    return rows

# 获取chrome handless 浏览器
browser = share_browser()
if __name__ == '__main__':
    # excel中的数据的 A1 取出来有个引号 从A2开始取数据
    # r = 2
    # query_list = ['115754-20-6']
    # 判断一下数据取完了没
    k = 0
    query_list=[]
    for r in range(find_excel_rows()-1):
        # 从 第二个开始取数据
        query_list.append(read_excel_file(r+1))
    #     把数据存到query_list中

    # 需要定制的请求头
    headers = {
        'referer': 'https://pubchem.ncbi.nlm.nih.gov/'
    }

    for query in query_list:
        k += 1
        # 下面这个是为了获取需要链接的cid号
        url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
            query) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
        response = requests.get(url=url, headers=headers)
        response.encoding = 'utf-8'
        # 下面这个json数据可以采集到
        json_content = response.text
        # print(page_content)
        # ***可以直接读取json，就不存进去了
        # with open('cid_3.json', 'w', encoding='utf-8') as fp:
        #             #     fp.write(str(json_content))
        #             # #   'https://pubchem.ncbi.nlm.nih.gov/compound/21180297'
        #             # obj = json.load(open('cid_3.json', 'r', encoding='utf-8'))
        obj = json.loads(json_content)
        obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
        # print(obj_cid)

        # src_url='https://pubchem.ncbi.nlm.nih.gov/compound/'+str(obj_cid)
        # # 下载链接所在的页面
        # response=requests.get(url=src_url,headers=headers)

        down_src = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/' + str(
            obj_cid) + '/record/SDF/?record_type=3d&response_type=save&response_basename=Conformer3D_CID_' + str(
            obj_cid)
        browser.get(down_src)
        time.sleep(4)
        print('第{0}条数据下载成功'.format(k))
        # time.sleep(0.5)
        # 关闭浏览器

        # print(down_src)
        # #把下载地址发送给requests模块
        # f=requests.get(down_src)
        # #下载文件
        # with open("cid.png","wb") as fp:
        #     fp.write(f.content)
# else:
#     browser.quit()
#     print('数据全部采集成功！')
