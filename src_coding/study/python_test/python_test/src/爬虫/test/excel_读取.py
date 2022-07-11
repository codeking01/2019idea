import openpyxl

def open_excel_file():
    wb=openpyxl.load_workbook('query.xlsx')
    k=3
    # print(type(wb))
    # sheets=wb.get_sheet_names
    # sheet=wb.get_sheet_by_name('SheetName') 这个语法已经过时了 要更换成 sheet=sheet['SheetName']
    sheet=wb['Sheet1']
    print(sheet.max_row)
    # 这个获取到的是一个cell对象
    cell=sheet['A'+str(k)]
    # cell_2=sheet.cell(row=1,column=1).value
    con=cell.value
    print(con)

    # print(cell_2)
    # print(type(cell_2))
    # cell=sheet.cell(row=2,column=2).value.split("'")
    # a="'"+cell.value
    # q_list=[a]
    # print(q_list[0])


open_excel_file()
