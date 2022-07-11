import json

import openpyxl
import scrapy
from scrapy import Request
from pubchem_test.items import PubchemTestItem


def read_excel_file():
    # r代表行数
    # 获取excle文件里面的内容
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    # print(type(wb))
    sheet = wb['Sheet1']
    return sheet

def read_excel_data(sheet,r):
    # 用来查看cas_data
    # sheet.cell(row=1,column=1).value
    # cell=sheet['A1']
    # **采集到要查询的数据 是str类型
    cell = sheet.cell(row=r, column=1).value
    # print(cell)
    if cell is not None:
        # 方便后期维护
        new_cell = cell
        # query_list = [new_cell]
        r += 1
        return new_cell
    else:
        return None

def find_excel_rows():
    # 用来查看表格的最大行数
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['Sheet1']
    # 获取到行数
    rows = sheet.max_row
    return rows


query_list = []
CAS_list=[]
# 用来记录爬取失败的cas号
pending_cas_list=[]

# 先从excle中把需要的数据全部取出来
cas_data=read_excel_file()
for r in range(100):   #find_excel_rows()-1
    # 从 第二个开始取数据,并且把数据存到query_list中，这r是从0开始的
    CAS_list.append(read_excel_data(cas_data,r + 2))  #这个是用来存储最后的名字
    # cas号传递给query去下载
    query_list.append(read_excel_data(cas_data,r + 2).replace('_','-')) #把传入的‘_’ 替换为‘-’



class PubSpider(scrapy.Spider):
    name = 'pub'
    allowed_domains = ['pubchem.ncbi.nlm.nih.gov']
    start_urls = ['https://pubchem.ncbi.nlm.nih.gov'] # It is invalid，I have overwrited the method of 'start_request'!
    def start_requests(self):

        index = 0
        # url='https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%2257-27-2%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
        for query in query_list:
            index += 1
            try:
                # 下面这个是为了获取需要链接的cid号
                url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
                    query) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
                # 根据url发请求
                print('第{0}条数据开始处理...'.format(index))
                cas=CAS_list[index-1]
                pending_cas=cas
                # pending_data=PubchemTestItem(pending_cas=pending_cas)
                # 把index、query、pending_cas_list都往下传，如果后面解析错误会进行记录
                yield scrapy.Request(url=url,callback=self.parse,meta={'cas':cas,'index':index,'query':query})

            except:
                cas=CAS_list[index-1]
                # 把未下载的cas号传递存储给pending_cas_list
                pending_cas_list.append(query)
                print('Warning:--第{0}条数据的cas号可能有问题...正在记录'.format(index),pending_cas_list)
                yield pending_cas_list

    def parse(self, response):
        try:
            query=response.meta['query']
            # index是用来保存爬取失败的数据的索引
            index=response.meta['index']
            cas=response.meta['cas']
            # response.encoding='utf-8' 这里不需要编码（应该是scrapy框架已经做了处理了）
            json_content = response.text
            obj = json.loads(json_content)
            # 找到json中相应的数据
            obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
            # print(obj_cid)    这个地方是根据cid号进行下载的，做了加密，直接解析是获取不到地址的
            down_src = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/' + str(
                obj_cid) + '/record/SDF/?record_type=3d&response_type=save&response_basename=Conformer3D_CID_' + str(
                obj_cid)
            pic=PubchemTestItem(down_src=down_src,obj_cid=obj_cid,cas=cas,index=index,query=query)
            yield pic
            print('第{0}条数据传递成功'.format(index),'**当前有问题的CAS号的列表：',pending_cas_list)

        except:
            pending_cas_list.append(query)
            print('Warning:！！第{0}条数据的cid号有问题...正在记录'.format(index))
            print('{0}在****parse下出问题'.format(cas),pending_cas_list)
            yield pending_cas_list