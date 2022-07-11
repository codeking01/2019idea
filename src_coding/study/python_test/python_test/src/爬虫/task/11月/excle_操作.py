import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
# # 读取excle档案
wb=load_workbook('new_excle.xlsx')
# # worksheet的缩写
# ws=wb.active
# # 创建工作表
# wb.create_sheet('新表测试')
# wb.save('new_excle.xlsx')

# # 找到工作表的名字
# # name=wb.sheetnames
# # print(name)
# # 指定操作的表
# ws=wb['Sheet1']


# 修改操作
# ws['C3'].value='hahhahh'
# 保存，只有保存了，上述的操作才有效
# wb.save('解放西路.xlsx')


# # 创建excle档案
# wb=Workbook()
# ws=wb.active
# ws.title='第二个exlce的新表'
# # ws['A1']=12313213
# # 新增一行数据
# # 传 那个excle的字母  比如ABCDE等
# 插入数据（横排）
# ws.insert_row(3)
# 插入列 B对应2  删除的话就改成delete_cols
# ws.insert_cols(2)
# ws.append(['店名','链接'])
# wb.save('new_excle.xlsx')

# wb=load_workbook('new_excle.xlsx')
# ws=wb.active
# char=get_column_letter(1)  #str类型
# print(type(char))
# print(ws[char+'1'])
# print(ws[char+'1'].value)


# 合并单元格
# ws.merge_cells('A1:E1')
# 取消合并单元格 原来的数据不会恢复
# ws.unmerge_cells('A1:E1')