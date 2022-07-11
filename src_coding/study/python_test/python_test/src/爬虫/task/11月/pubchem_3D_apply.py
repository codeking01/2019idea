import json
import openpyxl
import random
import requests
import urllib.request

...
# 原理：
#     根据excle中的给定的query号，通过调试隐藏接口，获取到存放JSON数据的接口，
#     根据解析json去获取到pubchem加密的cid号，然后在获取到下载到链接
#     下载方式：
#      1.通过chorme_handless去点击下载链接完成操作
#      2.直接通过urlretrieve去下载
# ******难点：1.找隐藏的接口，2.破解隐藏下载链接进行拼接
#       坑点：在A1数据位置主要随便写一个数进去，否则有bug(已经修复)
...


# 封装这个浏览器对象
# def share_browser():
#     chrome_options = Options()
#     # 如果该目录不存在则直接创建 下载的东西存在 D:\A_test_pic
#     prefs = {"download.default_directory": "D:\A_test_pic"}
#     # 将自定义设置添加到chrome配置对象实例中
#     chrome_options.add_experimental_option("prefs", prefs)
#
#     chrome_options.add_argument('‐‐headless')
#     chrome_options.add_argument('‐‐disable‐gpu')
#     path = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
#     chrome_options.binary_location = path
#     browser = webdriver.Chrome(chrome_options=chrome_options)
#     return browser

#
    # 2021年10月28日 增加数据库连接存储
#
import pymysql
# 打开数据库连接
db = pymysql.connect(host="localhost", user="root", password="123456", database="417database")
# 使用cursor（）方法获取游标
cursor=db.cursor()
# # 执行SQL语句
# insert_sql = "insert into test01(id, name, age) values(%s, %s, %s)"
# parm=(4,"是的",2)
# # 执行sql语句
# cursor.execute(insert_sql,parm)
# # 提交到数据库执行
# db.commit()
# cursor.execute("select * from test01")
# # 查看表里所有数据
# data = cursor.fetchall()
# print(data)



def read_excel_file():
    # r代表行数
    # 获取excle文件里面的内容
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    # print(type(wb))
    sheet = wb['Sheet1']
    return sheet

def read_excel_data(sheet, r):
    # 用来查看cas_data
    # sheet.cell(row=1,column=1).value
    # **采集到要查询的数据 是str类型
    cell = sheet.cell(row=r, column=1).value
    # print(cell)
    if cell is not None:
        # 方便后期维护
        new_cell = cell
        # query_list = [new_cell]
        r += 1
        return new_cell
    else:
        return None


def find_excel_rows():
    # 用来查看表格的最大行数
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['Sheet1']
    # 获取到行数
    rows = sheet.max_row
    return rows


query_list = []
CAS_list = []
# 用来记录爬取失败的CAS号，存到这个表，等待下次在爬取
pending_cas_list = []

# 先从excle中把需要的数据全部取出来
cas_data = read_excel_file()
for r in range(51,100):  # find_excel_rows()-1
    # 从 第二个开始取数据,并且把数据存到query_list中，这r是从0开始的,不包括尾巴
    CAS_list.append(read_excel_data(cas_data, r + 2))  # 这个是用来存储最后的名字
    query_list.append(read_excel_data(cas_data, r + 2).replace('_', '-'))  # 把传入的‘_’ 替换为‘-’

# 获取chrome handless 浏览器
# browser = share_browser()
if __name__ == '__main__':
    # excel中的数据的 A1 取出来有个引号 从A2开始取数据
    # r = 2
    # query_list = ['115754-20-6']
    # index用来计数
    index = 0
    # 需要定制的请求头
    headers = {
        'referer': 'https://pubchem.ncbi.nlm.nih.gov/'
    }
    # 加代理池
    proxies_pool = [
        {'http': '112.195.241.83: 3256'},
        {'http': '117.94.222.222: 3256'},
        {'http': '117.94.222.28	: 3256'},
        {'http': '121.230.211.227:3256'},
        {'http': '111.72.25.79:3256'},
        {'http': '47.75.132.50:8118'},
        {'http': '27.191.60.200:3256'},
        {'http': '219.151.142.29:3128'},
        {'http': '211.65.197.93:80'},
        {'http': '211.65.197.93:80'},
        {'http': '121.232.148.242:3256'},
    ]

    # 遍历这个query_list,开始相应的打印操作
    for query in query_list:
        index += 1
        try:
            cas = CAS_list[index - 1]
            # 下面这个是为了获取需要链接的cid号
            url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
                query) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
            proxies = random.choice(proxies_pool)
            response = requests.get(url=url, headers=headers, proxies=proxies)
            response.encoding = 'utf-8'
            # 下面这个json数据可以采集到
            json_content = response.text
            # print(page_content)
            # ***可以直接读取json，就不存进去了
            obj = json.loads(json_content)
            # 找到json中相应的数据
            obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
            # print(obj_cid)    这个地方是根据cid号进行下载的，做了加密，直接解析是获取不到地址的

            down_src = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/' + str(
                obj_cid) + '/record/SDF/?record_type=3d&response_type=save&response_basename=Conformer3D_CID_' + str(
                obj_cid)

            urllib.request.urlretrieve(down_src, './3D_pic/Conformer3D_CID_' + str(cas) + '.sdf')
            insert_sql = "insert into test(cas,image, src) values(%s,%s,%s)"
            fp = open("./3D_pic/Conformer3D_CID_"+str(cas)+'.sdf','rb')
            img = fp.read()
            fp.close()
            parm=(cas,img,down_src)
            cursor.execute(insert_sql,parm)
            db.commit()
            print('第{0}条数据下载成功'.format(index))

        except:
            insert_sql = "insert into failedlist(cas, src) values(%s, %s)"
            parm=(cas,down_src)
            cursor.execute(insert_sql,parm)
            db.commit()
            print('Waring：第{0}条数据下载失败.....正在记录，请稍后...'.format(index))
            pending_cas_list.append(query)


    # 把爬取失败的数据写入新的excle表格中
    # 定义指定输出Excel文件的名称，读入方式，编码方式
    fp = open('result.xls', 'w', encoding='gbk')
    # 参数'w'表示往指定表格读入数据，会先将表格中原本的内容清空
    # 若把参数’w'修改为‘a+',即可实现在原本内容的基础上，增加新写入的内容
    for i in range(0, len(pending_cas_list)):
        fp.write(str(pending_cas_list[i]))
        fp.write('\n')  # '\t'表示每写入一个元素后，会移动到同行的下一个单元格
    fp.write("\n")  # 换行操作
    fp.close()

# 用selenium实现的
# browser.get(down_src)
# time.sleep(4)
# print('第{0}条数据下载成功'.format(index))
# # time.sleep(0.5)
# 关闭浏览器
# browser.quit()
# 10024-74-5
