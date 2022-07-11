# 店名链接：http://www.dianping.com/shop/H8qpxHT6tl72i2ME
# 首页地址：http://www.dianping.com/changsha/ch10/g117
# 第二页地址：http://www.dianping.com/changsha/ch10/g117p2
import json
import random
import re
import time
import urllib
import urllib.request
from lxml import etree

import requests
import xlwt
from bs4 import BeautifulSoup

from openpyxl import Workbook,load_workbook

def create_request(page):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36',
        # 'Cookie': 'navCtgScroll=0; _lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ua=dpuser_15059620632; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; fspop=test; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1634227745,1635433785,1635434475; cy=344; cye=changsha; thirdtoken=41702c09-7086-430b-ab15-e7448c57e6c0; dper=234cb94f92704aef13a7e6c816ada0ec5d82dfce20785f8f7461236e5ad4452e080a2eb9d7089f0779c7dccdb93f453054a0e3dd272f57e46cf3f29951cb4adf5159c53995424956e2bae8b3aa6d15febd25ed5e382eabe35e43e4007f6a7b94; ll=7fd06e815b796be3df069dec7836c3df; ctu=b58a9fce079b27059480a633b27a83c81b97ef52648306836b7b2a7e61a47c0127548a04b35974cf992382a35b6cbb17; dplet=7a0ee364bff4bc2bf9ed9183621d7340; _lxsdk_s=17cc773261d-562-249-600%7C%7C710; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1635436153',
        # 'Cookie': '_lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; fspop=test; ua=sunshine; dper=234cb94f92704aef13a7e6c816ada0ec99322848b9a103cfe3bd94b17b5973abe3504b1c18ec4575f82a7f8e19bc2765a23a9a9541ab953eeca862f3c2712be82cce25199a025ec99a3cb8500c39eae6f33a8983370d8fb82c9739611bd81246; ll=7fd06e815b796be3df069dec7836c3df; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1635434475,1635479818,1635574037,1635611895; cye=shanghai; cy=1; dplet=d34cef477d103bc73ac23b9303221b73; _lxsdk_s=17cd192f0f0-4b9-470-d78%7C%7C740; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1635614015',
        # 'Cookie':'_lxsdk_cuid=17cd62cf030c8-052c48423c0223-b7a1438-100200-17cd62cf030c8; _lxsdk=17cd62cf030c8-052c48423c0223-b7a1438-100200-17cd62cf030c8; _hc.v=fe0e93b2-1d50-ef87-463d-72c879790c05.1635680843; _dp.ac.v=909faa0d-b5dd-410a-bffd-4fdf8602239d; UM_distinctid=17cd62d17474e3-046570cd0fcfb1-b7a1438-100200-17cd62d17483dc; _thirdu.c=65af27757b431c147ee940416f00e97e; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1635680843,1635680958; thirdtoken=b940c744-6159-41ec-b9c3-e5fb9679adec; dplet=f91ba0fb13b45054537af6aa06d2ff15; dper=31006a08db8bb685eba8e1e12a2bdf90538d135f04f50a06d507a7874ee0b659ed869cc2bb936d6fbd3617a0832d8bd2d43121ae6cd4f53c9614c8bf6defd66733828ae8511bfaf43a9710f0372232ac8eadf73b3e10e35f6e95926cf3f668a8; ll=7fd06e815b796be3df069dec7836c3df; ua=dpuser_7682168471; ctu=32aca2d1923ca96bb8b12d2ae498487616da8190723e0efa7e22be5d926fb20c; uamo=18522514966; fspop=test; cy=10; cye=tianjin; s_ViewType=10; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1635681816; _lxsdk_s=17cd62cf032-8a7-9ed-592%7C%7C338',
        'Cookie':'_lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; cye=changsha; aburl=1; Hm_lvt_dbeeb675516927da776beeb1d9802bd4=1635694823; Hm_lvt_4c4fc10949f0d691f3a2cc4ca5065397=1635694848; fspop=test; cy=344; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1636382190,1636449871,1636450245,1636450373; dplet=1a1121ffc4e10ebeecf02b3f0486ba5b; dper=55d202634e43c11ec7db27244d434825707374b6f6193fc1f03d7bbbc4443fd78818ab61e7394d8a5680082f910a96b6204b9ff737cdda3e44f8f52dbf2e75cd9f395abda79c1f158420ff4fa4c96312a03064aa861932533749ad11d774d165; ll=7fd06e815b796be3df069dec7836c3df; ua=sunshine; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1636450689; _lxsdk_s=17d04036412-835-490-167%7C%7C303',
        'Host': 'www.dianping.com',
        # 'Referer': 'http://www.dianping.com/changsha/ch10/g117',
        'Connection': 'keep-alive',
    }

    if page == 1:
        url = 'http://www.dianping.com/changsha/ch10/g112r301'
    else:
        url = 'http://www.dianping.com/changsha/ch10/g112r301p' + str(page)
    request = urllib.request.Request(url=url, headers=headers)
    return request


# http://www.dianping.com/shop/l3vlLMSKn6sZdJyy

def get_content(request):
    proxies_pool = [
        {'http': '139.198.121.76:8811'},
        {'http': '43.228.180.60:80'},
        {'http': '183.236.232.160:8080'},
        # {'http': '49.89.222.124:7082'},
        # {'http': '61.216.156.222:60808'},
        # {'http': '101.133.152.90:7788'},
        # {'http': '140.249.185.87:6969'},
        # {'http': '8.136.6.248：7788'},
        # {'http': '112.81.43.216:8888'},
        # {'http': '106.55.148.166:7890'},
        # {'http': '49.85.112.173:7890'},
    ]
    proxies = random.choice(proxies_pool)
    handler = urllib.request.ProxyHandler(proxies=proxies)
    opener = urllib.request.build_opener(handler)
    # response = urllib.request.urlopen(request)
    response = opener.open(request)
    content = response.read().decode('utf-8')
    return content
# 利用正则找到合适的需要的地址
# def get_src(url):
#     pattern = 'http://www.dianping.com/shop/+.*'
#     p = re.compile(pattern)
#     match = re.search(p, url)
#     if match is not None:
#         return match.group()
#     else:
#         return None

def down_load(content):
    global k
    # proxies_pool = [
    #     {'http': '140.249.48.241:6969'},
    #     {'http': '114.115.211.214:8085'},
    #     {'http': '175.153.233.52:8118'},
    #     {'http': '49.89.222.124:7082'},
    #     {'http': '61.216.156.222:60808'},
    #     {'http': '101.133.152.90:7788'},
    #     {'http': '140.249.185.876969'},
    #     {'http': '8.136.6.248:7788'},
    #     {'http': '112.81.43.216:8888'},
    #     {'http': '106.55.148.166:7890'},
    #     {'http': '49.85.112.173:7890'},
    # ]
    # proxies = random.choice(proxies_pool)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4595.0 Safari/537.36',
        'Cookie':'navCtgScroll=0; showNav=#nav-tab|0|1; navCtgScroll=0; _lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; cye=changsha; aburl=1; Hm_lvt_dbeeb675516927da776beeb1d9802bd4=1635694823; Hm_lvt_4c4fc10949f0d691f3a2cc4ca5065397=1635694848; fspop=test; cy=344; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1636382190,1636449871,1636450245,1636450373; dplet=1a1121ffc4e10ebeecf02b3f0486ba5b; dper=55d202634e43c11ec7db27244d434825707374b6f6193fc1f03d7bbbc4443fd78818ab61e7394d8a5680082f910a96b6204b9ff737cdda3e44f8f52dbf2e75cd9f395abda79c1f158420ff4fa4c96312a03064aa861932533749ad11d774d165; ll=7fd06e815b796be3df069dec7836c3df; ua=sunshine; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1636463808; _lxsdk_s=17d04d1a9c1-d8e-097-441%7C%7C168',
        'Host': 'www.dianping.com',
        # 'Referer': 'http://www.dianping.com/changsha/ch10/g112',
        'Connection': 'keep-alive',
    }
    proxies_pool = [
        {'http': '139.198.121.76:8811'},
        {'http': '43.228.180.60:80'},
        {'http': '183.236.232.160:8080'},
        # {'http': '49.89.222.124:7082'},
        # {'http': '61.216.156.222:60808'},
        # {'http': '101.133.152.90:7788'},
        # {'http': '140.249.185.87:6969'},
        # {'http': '8.136.6.248：7788'},
        # {'http': '112.81.43.216:8888'},
        # {'http': '106.55.148.166:7890'},
        # {'http': '49.85.112.173:7890'},
    ]
    proxies = random.choice(proxies_pool)
    # 解析数据
    tree = etree.HTML(content)
    # 解析想要的指定的内容
    name_list = tree.xpath('//div[@class="txt"]/div[@class="tit"]/a/h4/text()')
    src_list = tree.xpath('//div[@class="txt"]/div[@class="tit"]/a[1]/@href')
    # url = 'http://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=H9Vnbc8KMYG56vZN&cityId=344&mainCategoryId=34246'
    # url = 'http://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=laHrfapSRgeinxn1&cityId=344&mainCategoryId=34245'
    # https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=l6owxKl2JjGece58&cityId=344&mainCategoryId=34245
    # print(pic_list)
    # print(name_list)

    for i in range(len(src_list)):
        # 店的网址
        src = src_list[i]
        # print(src)
        # if get_src(src) is not None:
        k += 1
        print(name_list[i])
        try:
            # 进入每个页面以后获取需要的数据
            url = src
            # handler = urllib.request.ProxyHandler(proxies=proxies)
            # opener = urllib.request.build_opener(handler)
            print(url)
            response=requests.get(url=url,headers=headers,proxies=proxies).text

            # request = urllib.request.Request(url=url, headers=headers)
            # print(request)
            # response = opener.open(request)
            print(response)
            # response = urllib.request.urlopen(request)
            time.sleep(random.randint(2, 4))
            print('response请求成功')
            tree=etree.HTML(response)
            titles=tree.xpath('//body[@id="top"]/script[1]//text()')
            titles=titles[0]
            print(titles)

            # soup = BeautifulSoup(response.read().decode('utf-8'), 'lxml')
            # print(soup)
            # titles = str(soup.select('body[id="top"]>script', limit=1)[0])  # CSS 选择器  找到需要的字段
            # print(type(titles))

            # 找到shopCityId
            pattern = 'shopCityId: (.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match)
            pattern = '\d{1,8}'
            p = re.compile(pattern)
            match = re.search(p, str(match.group()))
            shopCityId = match.group()
            # print(shopCityId)
            # 获取shopId
            pattern = 'shopId: (.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match.group())
            shopId = str(match.group())
            shopId = shopId.split('"')[1]
            # print(shopId)
            # 获取mainCategoryId
            pattern = 'mainCategoryId:(.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match)
            pattern = '\d{1,8}'
            p = re.compile(pattern)
            match = re.search(p, str(match.group()))
            mainCategoryId = match.group()
            print('hhhhhh')
            # print(mainCategoryId)
            final_url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=' + str(
                shopId) + '&cityId=' + str(shopCityId) + '&mainCategoryId=' + str(mainCategoryId)
            # final_url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?'
            headers1={
                'Referer': 'http://www.dianping.com/shop/l9eLBvYIHEtYtvvy',
                'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36',
                'Cookie':'_lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; cye=changsha; aburl=1; Hm_lvt_dbeeb675516927da776beeb1d9802bd4=1635694823; Hm_lvt_4c4fc10949f0d691f3a2cc4ca5065397=1635694848; fspop=test; cy=344; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1636382190,1636449871,1636450245,1636450373; dplet=1a1121ffc4e10ebeecf02b3f0486ba5b; dper=55d202634e43c11ec7db27244d434825707374b6f6193fc1f03d7bbbc4443fd78818ab61e7394d8a5680082f910a96b6204b9ff737cdda3e44f8f52dbf2e75cd9f395abda79c1f158420ff4fa4c96312a03064aa861932533749ad11d774d165; ll=7fd06e815b796be3df069dec7836c3df; ua=sunshine; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1636450689; _lxsdk_s=17d04036412-835-490-167%7C%7C303',

            }
            parameters={
                'shopId':shopId,
                'cityId':shopCityId,
                'mainCategoryId':mainCategoryId,
                # '_token':'eJx1UE1vgkAQ/S9z3sDuuqxA4kHTqhBJLaJGGg8UEQiIyBLoR/rfO6T20EOTSebNe2/yJvMJjXMCm1FKBSPQJQ3YwDSqSSDQKlTkyJAmszjjhkUg/stZ5pjAa7N7APuFjaQkzKDHgfGR+GHG0jySO+QIucAaPA5aIGvb2tb1vu+1Ux5VdV6lWny96Cq71nppJatZd3CWj+2h7bp3POk/f5xFVaqyCAHVGy4o5YAhl2AIEdwiTJi4zpkkUgxKMSjYo3tvf2cPH4BOlacVosR9CzZKqNvZ91Sw5d5uuQriDy+mQm22o/Wqd9yFXxwus/w235+LeT81nsKoXhdGWYQiafK97+53Zbm9Tp3wOVikkwl8fQNf32uM',
                # 'uuid':'fe0e93b2-1d50-ef87-463d-72c879790c05.1635680843',
                # 'platform':'1',
                # 'partner':'150',
                # 'optimusCode':'10',
                # 'originUrl':'http://www.dianping.com/shop/l9eLBvYIHEtYtvvy'
            }

            print(final_url)



            # worksheet.write(k, 0, label=name_list[i])
            # worksheet.write(k, 1, label=final_url)
            # worksheet.write(k, 2, label=url)
            time.sleep(random.randint(4,5))

            # response1 = requests.get(url=final_url, headers=headers1,params=parameters)
            # print('response1......')
            # print(response1)
            # content = response1.text
            # # print(content)
            # dic = {'&#xe44a': '0', '&#xe3fc': '1', '&#xf5d6': '2', '&#xeedb': '3', '&#xe807': '4', '&#xe8c9': '5',
            #        '&#xe836': '6', '&#xf26a': '7', '&#xf447': '8', '&#xe405': '9', '.1': '1', '1': '1', }
            # # items=soup.find_all(class_='brief-info')
            # print('qqqqqqqq')
            # obj = json.loads(content)
            # print('obj进来了')
            # print(obj)
            # lenth = len(obj["avgPrice"].split('>'))
            # obj_price = obj["avgPrice"]
            # if (obj_price[0] == '0'):
            #     price = '--(没有)';
            # else:
            #     # 处理价格
            #     if (lenth == 3):
            #         b = obj_price.split('>')[1].split(';')[0]
            #         b = dic[b]
            #         # 要加‘1’  注意这个没有处理单价为1元
            #         if (obj_price[0] == '1'):
            #             if (obj_price[-1] == '1'):  # 处理101的情况
            #                 price = '1' + b + '1' + '元'
            #             # 处理110
            #             elif (obj_price[1] == '1'):
            #                 price = '11' + b + '元'
            #             else:
            #                 price = '1' + b + '元'
            #         elif (obj_price[-1] == '1'):  # 处理011
            #             if (obj_price[-2] == '1'):
            #                 price = b + '11' + '元'
            #             else:
            #                 price = b + '1' + '元'
            #         else:  # 处理个位数情况
            #             price = b + '元'
            #
            #     elif (lenth == 5):
            #         # 处理221  212 22 122
            #         # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xf5d6;</d', '']
            #         a = obj_price.split('>')[1].split(';')[0]
            #         a = dic[a]
            #         b = obj_price.split('>')[3].split(';')[0]
            #         b = dic[b]
            #         # print(obj_price.split('>'))
            #         # print(obj_price.split('>')[0][0])
            #         # 122
            #         if (obj_price.split('>')[0][0] == '1'):
            #             price = '1' + a + b + '元'
            #             # print('11111')
            #         # 221
            #         elif (obj_price.split('>')[2][0] == '1'):
            #             price = a + b + '1' + '元'
            #             # print('2222222')
            #         # 212
            #         elif (obj_price.split('>')[4] == '1'):
            #             price = a + '1' + b + '元'
            #             # print('3333333')
            #         else:
            #             # print('44444')
            #             price = a + b + '元'
            #     elif (lenth == 7):
            #         # 处理000的情况
            #         # o='<d class=\"num\">&#xf5d6;</d><d class=\"num\">&#xeedb;</d><d class=\"num\">&#xf26a;</d>'
            #         # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xeedb;</d', '<d class="num"', '&#xf26a;</d', '']
            #         a = obj_price.split('>')[1].split(';')[0]
            #         a = dic[a]
            #         b = obj_price.split('>')[3].split(';')[0]
            #         b = dic[b]
            #         c = obj_price.split('>')[5].split(';')[0]
            #         c = dic[c]
            #         price = a + b + c + '元'
            #     else:
            #         price = '--(没有)'
            # print(price)
            # # 评分
            # if ('fiveScore' not in obj.keys()):
            #     obj_sore = '--(没有)'
            # else:
            #     obj_sore = obj["fiveScore"]
            # # 口味
            # if (obj["shopRefinedScoreValueList"][0] == '-'):
            #     obj_taste = '--(没有)'
            # else:
            #     obj_taste_1 = obj["shopRefinedScoreValueList"][0].split('>')[1].split(';')[0]
            #     obj_taste_1 = dic[obj_taste_1]
            #     # obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
            #     obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')
            #     if (len(obj_taste_2) == 3):
            #         obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[2]
            #     else:
            #         obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
            #     obj_taste_2 = dic[obj_taste_2]
            #     obj_taste = obj_taste_1 + '.' + obj_taste_2
            #
            # # 环境
            # if (obj["shopRefinedScoreValueList"][1] == '-'):
            #     obj_env = '--(没有)'
            # else:
            #     # 环境
            #     obj_env_1 = obj["shopRefinedScoreValueList"][1].split('>')[1].split(';')[0]
            #     obj_env_1 = dic[obj_env_1]
            #     # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
            #     # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[2]
            #     obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')
            #     if (len(obj_env_2) == 3):
            #         obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[2]
            #     else:
            #         obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
            #     # print(obj_env_2)
            #     obj_env_2 = dic[obj_env_2]
            #     obj_env = obj_env_1 + '.' + obj_env_2
            # if (obj["shopRefinedScoreValueList"][2] == '-'):
            #     obj_serve = '--(没有)'
            # else:
            #     # 服务
            #     obj_serve_1 = obj["shopRefinedScoreValueList"][2].split('>')[1].split(';')[0]
            #     obj_serve_1 = dic[obj_serve_1]
            #     # obj_serve_2=obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
            #     obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')
            #     if (len(obj_serve_2) == 3):
            #         obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[2]
            #     else:
            #         obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
            #     # print(obj_serve_2)
            #     obj_serve_2 = dic[obj_serve_2]
            #     # print(obj_serve_2)
            #     obj_serve = obj_serve_1 + '.' + obj_serve_2
            # print(obj_sore, price, obj_taste, obj_env, obj_serve)
            #
            # worksheet.write(k, 0, label=name_list[i])
            # worksheet.write(k, 1, label=obj_sore)
            # worksheet.write(k, 2, label=price)
            # worksheet.write(k, 3, label=obj_taste)
            # worksheet.write(k, 4, label=obj_env)
            # worksheet.write(k, 5, label=obj_serve)
            # worksheet.write(k, 6, label=src)
            #
            # # request=urllib.request.Request(url=src,headers=headers)
            # # content=get_content(request)
            # # tree=etree.HTML(content)
            # # //div[@class="mid-score score-45"]
            # # mark_list=tree.xpath('//div[@class="star-wrapper"]/div')
            # # 评分地址http://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=H9Vnbc8KMYG56vZN&cityId=344&mainCategoryId=34246&_token=eJx1kM1OwkAUhd9l1pN2%2FilNXEA0UKBqSqkUw6IdsCXQUpgKqPHdvaO4cOHqfHPm3NyT%2B4GOwQr5lBAiKEan9RH5iDrEUQij1sCP4lJwj1DGGMdI%2F%2FWUHcqPyS3yn6nkCnMql9aJwPhxPCWW%2BIoMkAlsBeUBRFDZto3vuufz2VltsrrZ1IWj95Vryn3jDrtJnWtvHKYDqU6Le6j0X16XWV2YMgOgxC0o7SDYUcV2B%2Bt2oBeFaeZ5mEnxTRJqMEsdhmmX2%2FjWxkGzq7a%2F7xCOAlGzKWqg9egST40wh5coNPGMhclwEuv3UBNhpjP%2BOBFBMFhU6SFKatnXvYdIPaXz3egtytenZvjaVIvdbC7HvT5jd9v0Bn1%2BAc%2BGbCo%3D&uuid=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745&platform=1&partner=150&optimusCode=10&originUrl=http%3A%2F%2Fwww.dianping.com%2Fshop%2FH9Vnbc8KMYG56vZN
            wb.save('小吃快餐天心区_地址2.xlsx')
            print('第{0}条数据成功记录'.format(k))
            print()
        except:
            wb.save('小吃快餐天心区_地址2.xlsx')
            print('第{0}条数据处理失败'.format(k))
            print(src, i, name_list[i])
            print()


# "1<d class=\"num\">&#xe405;</d>"  1开头
# http://www.dianping.com/shop/H8qpxHT6tl72i2ME 单独处理 107


# else:
#     pass
# http://www.dianping.com/shop/G6cjEzJiVFGuY8Vm
# http://www.dianping.com/shop/FujE6yfc4atZpm5K
# if __name__ == '__main__':
k = 0
start_page = (int(input("请输入起始页码")))
end_page = (int(input("请输入结束页码")))

wb=Workbook()
ws=wb.active
ws.title='小吃快餐'

# workbook = xlwt.Workbook()
# 创建一个worksheet
# sheet_name=clib_name
# worksheet = workbook.add_sheet('中山亭_乐和城')
# worksheet1 = workbook.add_sheet('哈哈哈')
# worksheet.write(0, 0, label='店名')
# worksheet.write(0, 1, label='星级')
# worksheet.write(0, 2, label='人均')
# worksheet.write(0, 3, label='口味')
# worksheet.write(0, 4, label='环境')
# worksheet.write(0, 5, label='服务')
# worksheet.write(0, 6, label='网址查询')

for page in range(start_page, end_page + 1):
    #       获取定制对象
    request = create_request(page)
    # print(request)
    # 响应服务器数据
    content = get_content(request)
    # print(content)
    # 解析数据，并且下载
    down_load(content)

