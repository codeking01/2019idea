#
# import urllib.request
# #  123_3_45
# import openpyxl
#
# def read_excel_file(r):
#     # r代表行数
#     # 获取excle文件里面的内容
#     wb = openpyxl.load_workbook('query.xlsx')  #只有两个数据
#     # print(type(wb))
#     sheet = wb['Sheet1']
#     # sheet.cell(row=1,column=1).value
#     # cell=sheet['A1']
#     # **采集到要查询的数据 是str类型
#     cell = sheet.cell(row=r, column=1).value
#     # print(cell)
#     if cell != None:
#         # 方便后期维护
#         new_cell = cell
#         # query_list = [new_cell]
#         r += 1
#         return new_cell
#     else:
#         return None
#
#
# def find_excel_rows():
#     wb = openpyxl.load_workbook('query.xlsx')
#     sheet = wb['Sheet1']
#     # 获取到行数
#     rows = sheet.max_row
#     return rows
#
# if __name__ == '__main__':
#     print(find_excel_rows())
#     print('*********')
#     name_list=[]
#     for r in range(find_excel_rows()-1):
#         # 从 第二个开始取数据,并且把数据存到query_list中，这r是从0开始的
#         name_list.append(read_excel_file(r + 2).replace('-','_'))
#         print(len(name_list))
#         print('--') #里面是字符串类型的（str）
#         # name_list[r] = name_list[r].replace('-','_')
#         # k=0 #用来修改字符串 '_'
#         # for i in name_list[r]:
#         #     if i is '_':
#         #         i='-'
#
#     print(name_list)
#     print(name_list[0][2])


'''
# ------ 测试代码
import openpyxl

def read_excel_file():
    # r代表行数
    # 获取excle文件里面的内容
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    # print(type(wb))
    sheet = wb['Sheet1']
    return sheet

def read_excel_data(sheet,r):
    # 用来查看cas_data
    # sheet.cell(row=1,column=1).value
    # cell=sheet['A1']
    # **采集到要查询的数据 是str类型
    cell = sheet.cell(row=r, column=1).value
    # print(cell)
    if cell is not None:
        # 方便后期维护
        new_cell = cell
        # query_list = [new_cell]
        r += 1
        return new_cell
    else:
        return None

def find_excel_rows():
    # 用来查看表格的最大行数
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['Sheet1']
    # 获取到行数
    rows = sheet.max_row
    return rows


query_list = []
CAS_list=[]

# 先从excle中把需要的数据全部取出来
cas_data=read_excel_file()
for r in range(find_excel_rows()-1):
    # 从 第二个开始取数据,并且把数据存到query_list中，这r是从0开始的
    CAS_list.append(read_excel_data(cas_data,r + 2))  #这个是用来存储最后的名字
    query_list.append(read_excel_data(cas_data,r + 2).replace('_','-')) #把传入的‘_’ 替换为‘-’
print(CAS_list)
print(query_list)

'''
a=['1','2']
b=['1111','2']

def fun():
    print('hh')
    yield a
    print('1111')
    yield b
c=fun()
c.__next__()
d=c.__next__()
print(type(d))
print(d[0])