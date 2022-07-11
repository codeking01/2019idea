import openpyxl


def read_excel_file():
    # r代表行数
    # 获取excle文件里面的内容
    wb = openpyxl.load_workbook('长沙市五一广场.xlsx') #下次改成xlsx
    # print(type(wb))
    sheet = wb['五一广场']
    return sheet

def read_excel_data(sheet, r):
    # 用来查看cas_data
    # sheet.cell(row=1,column=1).value
    # **采集到要查询的数据 是str类型
    cell = sheet.cell(row=r, column=2).value
    # print(cell)
    if cell is not None:
        # 方便后期维护
        new_cell = cell
        # query_list = [new_cell]
        r += 1
        return new_cell
    else:
        return None


def find_excel_rows():
    # 用来查看表格的最大行数
    wb = openpyxl.load_workbook('长沙市五一广场.xlsx')
    sheet = wb['五一广场']
    # 获取到行数
    rows = sheet.max_row
    return rows
final_content=read_excel_file()
print(final_content)
print(find_excel_rows())
a=read_excel_data(final_content,1)
print(a)

