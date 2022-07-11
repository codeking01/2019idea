
import json
import random
import re
import time
from lxml import etree
from bs4 import BeautifulSoup

from openpyxl import Workbook
from selenium.webdriver import ActionChains

import requests

from selenium.webdriver.chrome.options import Options
from selenium import webdriver


def share_browser():
    chrome_options = Options()
    chrome_options.add_argument('‐‐headless')
    chrome_options.add_argument('‐‐disable‐gpu')
    UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36'
    chrome_options.add_argument(f'user-agent={UA}')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    path = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
    chrome_options.binary_location = path
    browser = webdriver.Chrome(chrome_options=chrome_options)
    return browser


# browser=share_browser()
# browser.maximize_window()
# # 去掉window.navigator.webdriver字段，防止被检测出是使用selenium
# browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
#     "source": """
#     Object.defineProperty(navigator, 'webdriver', {
#       get: () => undefined
#     })
#   """
# })

def login():
    """
    登录函数
    """
    login_url = 'https://account.dianping.com/login?redir=https://www.dianping.com'
    browser.get(login_url)  # 打开登录页面，手动扫码登录
    # while True:
    #     try:
    #         browser.find_element_by_class_name('login-container J-login-container')  # 还能找到"请登录"
    #         time.sleep(5)
    #     except:
    #         # browser.find_element_by_class_name('userinfo-container')  # 登录成功
    #         break
    print('登录成功')


def process_cookie(url):
    """
    处理cookie，将driver 的 cookie复制到 requests，并把cookie保存下来供下次使用
    :param refresh_cookie: 是否从selenium更新cookie
    :return:
    """
    headers = {
        'User-Agent': UA,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
    }
    session.headers.update(headers)
    # 将selenium中的cookie复制到requests.session()中
    # if refresh_cookie:
    # browser.refresh()
    browser.get(url)

    time.sleep(10)
    cookie = {cookie['name']: cookie['value'] for cookie in browser.get_cookies()}
    # 将获取到的cookie保存下来，避免重复获取
    with open('01cookies.txt', 'w') as f:
        f.write(json.dumps(browser.get_cookies()))
    #     # browser.quit()
    print('从selenium更新cookie')
    # else:
    with open('01cookies.txt', 'r') as f:
        cookies = json.loads(f.read())
        cookie = {_cookie['name']: _cookie['value'] for _cookie in cookies}
    print('从本地缓存更新cookie')
    session.cookies.update(cookie)


def crapy_data(url):
    global k
    print('---')
    # response = session.get(url,timeout=20).text

    browser.get(url)
    content = browser.page_source
    time.sleep(5)
    tree = etree.HTML(content)
    # 解析想要的指定的内容  拿到整页的标题和链接
    name_list = tree.xpath('//div[@class="txt"]/div[@class="tit"]/a/h4/text()')
    src_list = tree.xpath('//div[@class="txt"]/div[@class="tit"]/a[1]/@href')
    # 去把这一页的各个店子的数据取到
    for i in range(len(src_list)):
        k += 1
        # 店的网址
        src = src_list[i]
        print(name_list[i], src)
        try:
            # process_cookie(src)
            # 进入每个页面以后获取需要的数据
            # print(src)
            response = session.get(url=src, timeout=20)
            content = response.text
            # 用selenium操作
            # browser.get(url=src)
            # content=browser.page_source
            # time.sleep(random.randint(1, 3))

            print(content)
            tree = etree.HTML(content)
            shopId = src.split('shop/')[1]

            # 好评 中评 差评
            # good_nums=tree.xpath('//label[@class="filter-item J-filter-good"]/span[@class="count"]/text()')
            # mid_nums=tree.xpath('//label[@class="filter-item J-filter-common"]/span[@class="count"]/text()')
            # bad_nums=tree.xpath('//label[@class="filter-item J-filter-bad"]/span[@class="count"]/text()')
            # print(good_nums,mid_nums,bad_nums)

            # 取不到下面的数据了   下面的列表是空列表  2021.11.18
            titles = tree.xpath('//body[@id="top"]/script[1]//text()')
            titles = titles[0]
            print(titles)
            # 找到shopCityId
            pattern = 'shopCityId: (.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match)
            pattern = '\d{1,8}'
            p = re.compile(pattern)
            match = re.search(p, str(match.group()))
            shopCityId = match.group()
            # print(shopCityId)
            # 获取shopId
            pattern = 'shopId: (.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match.group())
            shopId = str(match.group())
            shopId = shopId.split('"')[1]
            # print(shopId)
            # 获取mainCategoryId
            pattern = 'mainCategoryId:(.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match)
            pattern = '\d{1,8}'
            p = re.compile(pattern)
            match = re.search(p, str(match.group()))
            mainCategoryId = match.group()
            print('进来了')
            # print(mainCategoryId)
            # final_url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=' + str(
            #     shopId) + '&cityId=' + str(shopCityId) + '&mainCategoryId=' + str(mainCategoryId)
            final_url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=' + str(
                shopId) + '&cityId=' + str(344) + '&mainCategoryId=' + str(mainCategoryId)
            print(final_url)

            process_cookie(final_url)

            response = session.get(final_url, timeout=None)
            # browser.get(final_url)
            print('final_url的', response)
            content = response.text
            print(content)

            # 这个需要经常修改
            dic = {'&#xe2f9': '0', '&#xee1b': '1', '&#xf694': '2', '&#xee43': '3', '&#xef49': '4', '&#xe598': '5',
                   '&#xee7e': '6', '&#xed03': '7', '&#xefff': '8', '&#xe773': '9', }
            # items=soup.find_all(class_='brief-info')

            obj = json.loads(content)
            print('obj加载成功！')
            print(obj)
            lenth = len(obj["avgPrice"].split('>'))
            obj_price = obj["avgPrice"]
            if (obj_price[0] == '0'):
                price = '--(没有)'
            else:
                # 处理价格
                if (lenth == 3):
                    b = obj_price.split('>')[1].split(';')[0]
                    b = dic[b]
                    # 要加‘1’  注意这个没有处理单价为1元
                    if (obj_price[0] == '1'):
                        if (obj_price[-1] == '1'):  # 处理101的情况
                            price = '1' + b + '1' + '元'
                        # 处理110
                        elif (obj_price[1] == '1'):
                            price = '11' + b + '元'
                        else:
                            price = '1' + b + '元'
                    elif (obj_price[-1] == '1'):  # 处理011
                        if (obj_price[-2] == '1'):
                            price = b + '11' + '元'
                        else:
                            price = b + '1' + '元'
                    else:  # 处理个位数情况
                        price = b + '元'

                elif (lenth == 5):
                    # 处理221  212 22 122
                    # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xf5d6;</d', '']
                    a = obj_price.split('>')[1].split(';')[0]
                    a = dic[a]
                    b = obj_price.split('>')[3].split(';')[0]
                    b = dic[b]
                    # print(obj_price.split('>'))
                    # print(obj_price.split('>')[0][0])
                    # 122
                    if (obj_price.split('>')[0][0] == '1'):
                        price = '1' + a + b + '元'
                        # print('11111')
                    # 221
                    elif (obj_price.split('>')[2][0] == '1'):
                        price = a + b + '1' + '元'
                        # print('2222222')
                    # 212
                    elif (obj_price.split('>')[4] == '1'):
                        price = a + '1' + b + '元'
                        # print('3333333')
                    else:
                        # print('44444')
                        price = a + b + '元'
                elif (lenth == 7):
                    # 处理000的情况
                    # o='<d class=\"num\">&#xf5d6;</d><d class=\"num\">&#xeedb;</d><d class=\"num\">&#xf26a;</d>'
                    # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xeedb;</d', '<d class="num"', '&#xf26a;</d', '']
                    a = obj_price.split('>')[1].split(';')[0]
                    a = dic[a]
                    b = obj_price.split('>')[3].split(';')[0]
                    b = dic[b]
                    c = obj_price.split('>')[5].split(';')[0]
                    c = dic[c]
                    price = a + b + c + '元'
                else:
                    price = '--(没有)'
            print(price)
            # 评分
            if ('fiveScore' not in obj.keys()):
                obj_sore = '--(没有)'
            else:
                obj_sore = obj["fiveScore"]
            if (obj["shopRefinedScoreValueList"][0] == '-'):
                obj_taste = '--(没有)'
            else:
                # 口味
                obj_taste_1 = obj["shopRefinedScoreValueList"][0].split('>')[1].split(';')[0]
                obj_taste_1 = dic[obj_taste_1]
                # obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
                obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')
                if (len(obj_taste_2) == 3):
                    obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[2]
                else:
                    obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
                obj_taste_2 = dic[obj_taste_2]
                obj_taste = obj_taste_1 + '.' + obj_taste_2
            if (obj["shopRefinedScoreValueList"][1] == '-'):
                obj_env = '--(没有)'
            else:
                # 环境
                obj_env_1 = obj["shopRefinedScoreValueList"][1].split('>')[1].split(';')[0]
                obj_env_1 = dic[obj_env_1]
                # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
                # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[2]
                obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')
                if (len(obj_env_2) == 3):
                    obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[2]
                else:
                    obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
                # print(obj_env_2)
                obj_env_2 = dic[obj_env_2]
                obj_env = obj_env_1 + '.' + obj_env_2
            if (obj["shopRefinedScoreValueList"][2] == '-'):
                obj_serve = '--(没有)'
            else:
                # 服务
                obj_serve_1 = obj["shopRefinedScoreValueList"][2].split('>')[1].split(';')[0]
                obj_serve_1 = dic[obj_serve_1]
                # obj_serve_2=obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
                obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')
                if (len(obj_serve_2) == 3):
                    obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[2]
                else:
                    obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
                # print(obj_serve_2)
                obj_serve_2 = dic[obj_serve_2]
                # print(obj_serve_2)
                obj_serve = obj_serve_1 + '.' + obj_serve_2
            print(obj_sore, price, obj_taste, obj_env, obj_serve)

            ws['A' + str(k + 1)].value = name_list[i]
            ws['B' + str(k + 1)].value = obj_sore
            ws['C' + str(k + 1)].value = price
            ws['D' + str(k + 1)].value = obj_taste
            ws['E' + str(k + 1)].value = obj_env
            ws['F' + str(k + 1)].value = obj_serve
            ws['G' + str(k + 1)].value = src

            # ws['G'+str(i+2)].value=good_nums
            # ws['H'+str(i+2)].value=mid_nums
            # ws['I'+str(i+2)].value=bad_nums
            # ws['J'+str(i+2)].value=src

            print('第{0}条数据成功记录'.format(k))
            wb.save('小吃快餐天心区.xlsx')
            time.sleep(5)
            print()
        except:
            wb.save('小吃快餐天心区.xlsx')
            print(src)
            print('第{0}条数据处理失败'.format(k))
            time.sleep(3)


if __name__ == '__main__':
    # 处理要写入的数据
    wb = Workbook()
    ws = wb.active
    ws.title = '小吃快餐'
import json
import random
import re
import time
from lxml import etree
from openpyxl import Workbook
import requests

from selenium.webdriver.chrome.options import Options
from selenium import webdriver


def share_browser():
    chrome_options = Options()
    chrome_options.add_argument('‐‐headless')
    chrome_options.add_argument('‐‐disable‐gpu')
    UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36'
    chrome_options.add_argument(f'user-agent={UA}')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    path = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
    chrome_options.binary_location = path
    browser = webdriver.Chrome(chrome_options=chrome_options)
    return browser


# browser=share_browser()
# browser.maximize_window()
# # 去掉window.navigator.webdriver字段，防止被检测出是使用selenium
# browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
#     "source": """
#     Object.defineProperty(navigator, 'webdriver', {
#       get: () => undefined
#     })
#   """
# })

def login():
    """
    登录函数
    """
    login_url = 'https://account.dianping.com/login?redir=https://www.dianping.com'
    browser.get(login_url)  # 打开登录页面，手动扫码登录
    # while True:
    #     try:
    #         browser.find_element_by_class_name('login-container J-login-container')  # 还能找到"请登录"
    #         time.sleep(5)
    #     except:
    #         # browser.find_element_by_class_name('userinfo-container')  # 登录成功
    #         break
    print('登录成功')


def process_cookie(url):
    """
    处理cookie，将driver 的 cookie复制到 requests，并把cookie保存下来供下次使用
    :param refresh_cookie: 是否从selenium更新cookie
    :return:
    """
    headers = {
        'User-Agent': UA,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
    }
    session.headers.update(headers)
    # 将selenium中的cookie复制到requests.session()中
    # if refresh_cookie:
    # browser.refresh()
    browser.get(url)

    time.sleep(10)
    cookie = {cookie['name']: cookie['value'] for cookie in browser.get_cookies()}
    # 将获取到的cookie保存下来，避免重复获取
    with open('01cookies.txt', 'w') as f:
        f.write(json.dumps(browser.get_cookies()))
    #     # browser.quit()
    print('从selenium更新cookie')
    # else:
    with open('01cookies.txt', 'r') as f:
        cookies = json.loads(f.read())
        cookie = {_cookie['name']: _cookie['value'] for _cookie in cookies}
    print('从本地缓存更新cookie')
    session.cookies.update(cookie)


def crapy_data(url):
    global k
    print('---')
    # response = session.get(url,timeout=20).text

    browser.get(url)
    content = browser.page_source
    time.sleep(5)
    tree = etree.HTML(content)
    # 解析想要的指定的内容  拿到整页的标题和链接
    name_list = tree.xpath('//div[@class="txt"]/div[@class="tit"]/a/h4/text()')
    src_list = tree.xpath('//div[@class="txt"]/div[@class="tit"]/a[1]/@href')
    # 去把这一页的各个店子的数据取到
    for i in range(len(src_list)):
        k += 1
        # 店的网址
        src = src_list[i]
        print(name_list[i], src)
        try:
            # process_cookie(src)
            # 进入每个页面以后获取需要的数据
            # print(src)
            response = session.get(url=src, timeout=20)
            content = response.text
            # 用selenium操作
            # browser.get(url=src)
            # content=browser.page_source
            # time.sleep(random.randint(1, 3))

            print(content)
            tree = etree.HTML(content)
            shopId = src.split('shop/')[1]

            # 好评 中评 差评
            # good_nums=tree.xpath('//label[@class="filter-item J-filter-good"]/span[@class="count"]/text()')
            # mid_nums=tree.xpath('//label[@class="filter-item J-filter-common"]/span[@class="count"]/text()')
            # bad_nums=tree.xpath('//label[@class="filter-item J-filter-bad"]/span[@class="count"]/text()')
            # print(good_nums,mid_nums,bad_nums)

            # 取不到下面的数据了   下面的列表是空列表  2021.11.18
            titles = tree.xpath('//body[@id="top"]/script[1]//text()')
            titles = titles[0]
            print(titles)
            # 找到shopCityId
            pattern = 'shopCityId: (.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match)
            pattern = '\d{1,8}'
            p = re.compile(pattern)
            match = re.search(p, str(match.group()))
            shopCityId = match.group()
            # print(shopCityId)
            # 获取shopId
            pattern = 'shopId: (.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match.group())
            shopId = str(match.group())
            shopId = shopId.split('"')[1]
            # print(shopId)
            # 获取mainCategoryId
            pattern = 'mainCategoryId:(.*?),'
            p = re.compile(pattern)
            match = re.search(p, titles)
            # print(match)
            pattern = '\d{1,8}'
            p = re.compile(pattern)
            match = re.search(p, str(match.group()))
            mainCategoryId = match.group()
            print('进来了')
            # print(mainCategoryId)
            # final_url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=' + str(
            #     shopId) + '&cityId=' + str(shopCityId) + '&mainCategoryId=' + str(mainCategoryId)
            final_url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=' + str(
                shopId) + '&cityId=' + str(344) + '&mainCategoryId=' + str(mainCategoryId)
            print(final_url)

            process_cookie(final_url)

            response = session.get(final_url, timeout=None)
            # browser.get(final_url)
            print('final_url的', response)
            content = response.text
            print(content)

            # 这个需要经常修改
            dic = {'&#xe2f9': '0', '&#xee1b': '1', '&#xf694': '2', '&#xee43': '3', '&#xef49': '4', '&#xe598': '5',
                   '&#xee7e': '6', '&#xed03': '7', '&#xefff': '8', '&#xe773': '9', }
            # items=soup.find_all(class_='brief-info')

            obj = json.loads(content)
            print('obj加载成功！')
            print(obj)
            lenth = len(obj["avgPrice"].split('>'))
            obj_price = obj["avgPrice"]
            if (obj_price[0] == '0'):
                price = '--(没有)'
            else:
                # 处理价格
                if (lenth == 3):
                    b = obj_price.split('>')[1].split(';')[0]
                    b = dic[b]
                    # 要加‘1’  注意这个没有处理单价为1元
                    if (obj_price[0] == '1'):
                        if (obj_price[-1] == '1'):  # 处理101的情况
                            price = '1' + b + '1' + '元'
                        # 处理110
                        elif (obj_price[1] == '1'):
                            price = '11' + b + '元'
                        else:
                            price = '1' + b + '元'
                    elif (obj_price[-1] == '1'):  # 处理011
                        if (obj_price[-2] == '1'):
                            price = b + '11' + '元'
                        else:
                            price = b + '1' + '元'
                    else:  # 处理个位数情况
                        price = b + '元'

                elif (lenth == 5):
                    # 处理221  212 22 122
                    # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xf5d6;</d', '']
                    a = obj_price.split('>')[1].split(';')[0]
                    a = dic[a]
                    b = obj_price.split('>')[3].split(';')[0]
                    b = dic[b]
                    # print(obj_price.split('>'))
                    # print(obj_price.split('>')[0][0])
                    # 122
                    if (obj_price.split('>')[0][0] == '1'):
                        price = '1' + a + b + '元'
                        # print('11111')
                    # 221
                    elif (obj_price.split('>')[2][0] == '1'):
                        price = a + b + '1' + '元'
                        # print('2222222')
                    # 212
                    elif (obj_price.split('>')[4] == '1'):
                        price = a + '1' + b + '元'
                        # print('3333333')
                    else:
                        # print('44444')
                        price = a + b + '元'
                elif (lenth == 7):
                    # 处理000的情况
                    # o='<d class=\"num\">&#xf5d6;</d><d class=\"num\">&#xeedb;</d><d class=\"num\">&#xf26a;</d>'
                    # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xeedb;</d', '<d class="num"', '&#xf26a;</d', '']
                    a = obj_price.split('>')[1].split(';')[0]
                    a = dic[a]
                    b = obj_price.split('>')[3].split(';')[0]
                    b = dic[b]
                    c = obj_price.split('>')[5].split(';')[0]
                    c = dic[c]
                    price = a + b + c + '元'
                else:
                    price = '--(没有)'
            print(price)
            # 评分
            if ('fiveScore' not in obj.keys()):
                obj_sore = '--(没有)'
            else:
                obj_sore = obj["fiveScore"]
            if (obj["shopRefinedScoreValueList"][0] == '-'):
                obj_taste = '--(没有)'
            else:
                # 口味
                obj_taste_1 = obj["shopRefinedScoreValueList"][0].split('>')[1].split(';')[0]
                obj_taste_1 = dic[obj_taste_1]
                # obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
                obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')
                if (len(obj_taste_2) == 3):
                    obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[2]
                else:
                    obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
                obj_taste_2 = dic[obj_taste_2]
                obj_taste = obj_taste_1 + '.' + obj_taste_2
            if (obj["shopRefinedScoreValueList"][1] == '-'):
                obj_env = '--(没有)'
            else:
                # 环境
                obj_env_1 = obj["shopRefinedScoreValueList"][1].split('>')[1].split(';')[0]
                obj_env_1 = dic[obj_env_1]
                # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
                # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[2]
                obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')
                if (len(obj_env_2) == 3):
                    obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[2]
                else:
                    obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
                # print(obj_env_2)
                obj_env_2 = dic[obj_env_2]
                obj_env = obj_env_1 + '.' + obj_env_2
            if (obj["shopRefinedScoreValueList"][2] == '-'):
                obj_serve = '--(没有)'
            else:
                # 服务
                obj_serve_1 = obj["shopRefinedScoreValueList"][2].split('>')[1].split(';')[0]
                obj_serve_1 = dic[obj_serve_1]
                # obj_serve_2=obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
                obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')
                if (len(obj_serve_2) == 3):
                    obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[2]
                else:
                    obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
                # print(obj_serve_2)
                obj_serve_2 = dic[obj_serve_2]
                # print(obj_serve_2)
                obj_serve = obj_serve_1 + '.' + obj_serve_2
            print(obj_sore, price, obj_taste, obj_env, obj_serve)

            ws['A' + str(k + 1)].value = name_list[i]
            ws['B' + str(k + 1)].value = obj_sore
            ws['C' + str(k + 1)].value = price
            ws['D' + str(k + 1)].value = obj_taste
            ws['E' + str(k + 1)].value = obj_env
            ws['F' + str(k + 1)].value = obj_serve
            ws['G' + str(k + 1)].value = src

            # ws['G'+str(i+2)].value=good_nums
            # ws['H'+str(i+2)].value=mid_nums
            # ws['I'+str(i+2)].value=bad_nums
            # ws['J'+str(i+2)].value=src

            print('第{0}条数据成功记录'.format(k))
            wb.save('小吃快餐天心区.xlsx')
            time.sleep(5)
            print()
        except:
            wb.save('小吃快餐天心区.xlsx')
            print(src)
            print('第{0}条数据处理失败'.format(k))
            time.sleep(3)


if __name__ == '__main__':
    # 处理要写入的数据
    wb = Workbook()
    ws = wb.active
    ws.title = '小吃快餐'
    ws['A1'] = '店名'
    ws['B1'] = '星级'
    ws['C1'] = '人均'
    ws['D1'] = '口味'
    ws['E1'] = '环境'
    ws['F1'] = '服务'
    ws['G1'] = '网址查询'

    # ws['G1']='好评数'
    # ws['H1']='中评数'
    # ws['I1']='差评数'
    # ws['J1']='网址查询'

    # refresh_cookie = True
    UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36'
    session = requests.session()
    session.headers.clear()

    # if refresh_cookie:
    browser = share_browser()
    browser.maximize_window()
    # 去掉window.navigator.webdriver字段，防止被检测出是使用selenium
    browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
        Object.defineProperty(navigator, 'webdriver', {
          get: () => undefined
        })
      """
    })
    login()
    # 更新cookie
    time.sleep(10)

    k = 0
    # 处理第一页的数据
    page = 1
    url = 'http://www.dianping.com/changsha/ch10/g112r301'
    # process_cookie(url)
    crapy_data(url)
    print('*******')
    time.sleep(2)

    # 处理第二页及以后的数据
    page += 1
    while (page > 1 and page < 51):  # 到最大50页为止
        url = 'http://www.dianping.com/changsha/ch10/g112r301p' + str(page)
        # process_cookie(url)
        # 爬取数据
        crapy_data(url)
        time.sleep(10)
        page += 1
    print('GG,全部爬取成功！！')

    # browser.get('http://www.dianping.com/shop/H2EXdyIsbUOcNXHy')
    # time.sleep(3)
    # a=browser.find_element_by_xpath('//span[@id="comment_score"]')
    # print(a.get_attribute('span'))
    # browser.quit()

    ws['A1'] = '店名'
    ws['B1'] = '星级'
    ws['C1'] = '人均'
    ws['D1'] = '口味'
    ws['E1'] = '环境'
    ws['F1'] = '服务'
    ws['G1'] = '网址查询'

    # ws['G1']='好评数'
    # ws['H1']='中评数'
    # ws['I1']='差评数'
    # ws['J1']='网址查询'

    # refresh_cookie = True
    UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36'
    session = requests.session()
    session.headers.clear()

    # if refresh_cookie:
    browser = share_browser()
    browser.maximize_window()
    # 去掉window.navigator.webdriver字段，防止被检测出是使用selenium
    browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
        Object.defineProperty(navigator, 'webdriver', {
          get: () => undefined
        })
      """
    })
    login()
    # 更新cookie
    time.sleep(10)

    k = 0
    # 处理第一页的数据
    page = 1
    url = 'http://www.dianping.com/changsha/ch10/g112r301'
    # process_cookie(url)
    crapy_data(url)
    print('*******')
    time.sleep(2)

    # 处理第二页及以后的数据
    page += 1
    while (page > 1 and page < 51):  # 到最大50页为止
        url = 'http://www.dianping.com/changsha/ch10/g112r301p' + str(page)
        # process_cookie(url)
        # 爬取数据
        crapy_data(url)
        time.sleep(10)
        page += 1
    print('GG,全部爬取成功！！')
