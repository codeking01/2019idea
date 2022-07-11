# url='https://www.chemicalbook.com/ProductChemicalPropertiesCB58227907.htm'
import random

import openpyxl
from openpyxl import *
from lxml import etree
import requests
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import time
from time import sleep


# def share_browser():
#     chrome_options = Options()
#     chrome_options.add_argument('‐‐headless')
#     chrome_options.add_argument('‐‐disable‐gpu')
#     UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36'
#     chrome_options.add_argument(f'user-agent={UA}')
#     chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
#     chrome_options.add_experimental_option('useAutomationExtension', False)
#
#     path = r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'
#     chrome_options.binary_location = path
#     browser = webdriver.Chrome(chrome_options=chrome_options)
#     return browser


# browser=share_browser()
# 创建Excel文件
wbk=Workbook()
ws=wbk.active
ws.title='cas_沸点'
ws['A1']='cas'
ws['B1']='沸点'

wb = openpyxl.load_workbook('Tb1.xlsx')
sheet = wb['data']
rows = sheet.max_row
tag=0
proxies_pool = [
    {'http': '8.218.81.68:59394'},
    {'http': '116.49.163.226:8080'},
    {'http': '1.15.88.109:7890'},
    {'http': '121.13.252.61:41564'},
    {'http': '111.231.86.149:7890'},
    {'http': '114.233.71.82:9000'},
    {'http': '202.109.157.60:9000'},
    {'http': '60.250.159.191:45983'},
    {'http': '121.13.252.61:41564'},
    {'http': '106.75.252.120:80'},
    {'http': '8.218.81.68:59394'},
]

proxies = random.choice(proxies_pool)
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'
}
com_point=0

# 起始页面 https://www.chemicalbook.com/Search.aspx?_s=&keyword=624-92-0
# 从2开始的 需要+1 每次
for k in range(9891,rows+1):  #下次 从1000,4000  记得修改第一个数字2050的数字有问题  4000-8000  4000可能需要修改 6356  8456  9891
    #处理遗漏的数据
    if(tag==1):
        t=k-1
        # 加个判断如果是和前一个一样的，直接写入
        com_cas=sheet.cell(row=t,column=2).value
        if(com_cas==sheet.cell(row=t-1,column=2).value):
            com_cas=com_cas.replace('_', '-')
            ws['A'+str(t)].value=com_cas
            ws['B'+str(t)].value=com_point
            wbk.save('cas_沸点1.xlsx')
            tag=0
            print('第{0}条记录成功'.format(t))
        else:
            cas =sheet.cell(row=t, column=2).value
            cas=cas.replace('_', '-')
            try :
                init_url = 'https://www.chemicalbook.com/Search.aspx?keyword=' + str(cas)
                print(cas,init_url)
                content = requests.get(init_url,headers=headers,timeout=10,proxies=proxies).text
                tree = etree.HTML(content)
                CBNumber=tree.xpath('//td[@data-field="cbnumber"]/text()')[0]
                print(CBNumber)
                # //*[@id="mbox"]/tbody/tr[1]/td[2]//text()
                url = 'https://www.chemicalbook.com/ProductChemicalProperties'+str(CBNumber)+'.htm'
                content = requests.get(url,headers=headers,timeout=10,proxies=proxies).text
                # print(content)
                tree = etree.HTML(content)
                a = tree.xpath('//tr[@class="ProdSupplierGN_ProductA_1"]/td//text()')
                b = tree.xpath('//tr[@class="ProdSupplierGN_ProductA_2"]/td//text()')
                point = '无'
                if (a is not None):
                    for i in range(len(a)):
                        if (a[i][0] == '沸'):
                            point = a[i + 1]
                            break
                        else:
                            pass
                if (b is not None):
                    for i in range(len(b)):
                        if (b[i][0] == '沸'):
                            point = b[i + 1]
                            break
                        else:
                            pass
                com_point=point
                ws['A'+str(t)].value=cas
                ws['B'+str(t)].value=point
                wbk.save('cas_沸点1.xlsx')
                tag=0
                print('第{0}条记录成功'.format(t))
            except:
                ws['A'+str(t)].value=cas
                ws['B'+str(t)].value='cas号可能有问题，请手动查询'
                wbk.save('cas_沸点1.xlsx')
                tag=0
                print('第{0}条记录失败'.format(t))

    # 处理重复数据
    com_cas=sheet.cell(row=k,column=2).value
    if(com_cas==sheet.cell(row=k-1,column=2).value):
        com_cas=com_cas.replace('_', '-')
        ws['A'+str(k)].value=com_cas
        ws['B'+str(k)].value=com_point
        wbk.save('cas_沸点1.xlsx')
        tag=0
        print('第{0}条记录成功'.format(k))
    else:
        cas =sheet.cell(row=k, column=2).value
        cas=cas.replace('_', '-')
        init_url = 'https://www.chemicalbook.com/Search.aspx?keyword=' + str(cas)
        print(cas,init_url)
        try:
            content = requests.get(init_url,headers=headers,timeout=10,proxies=proxies).text
            tree = etree.HTML(content)
            CBNumber=tree.xpath('//td[@data-field="cbnumber"]/text()')[0]
            print(CBNumber)
            # //*[@id="mbox"]/tbody/tr[1]/td[2]//text()

            url = 'https://www.chemicalbook.com/ProductChemicalProperties'+str(CBNumber)+'.htm'
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'

            }
            content = requests.get(url,headers=headers,timeout=10,proxies=proxies).text
            # print(content)
            tree = etree.HTML(content)
            a = tree.xpath('//tr[@class="ProdSupplierGN_ProductA_1"]/td//text()')
            b = tree.xpath('//tr[@class="ProdSupplierGN_ProductA_2"]/td//text()')
            point = '无'
            if (a is not None):
                for i in range(len(a)):
                    if (a[i][0] == '沸'):
                        point = a[i + 1]
                        break
                    else:
                        pass
            if (b is not None):
                for i in range(len(b)):
                    if (b[i][0] == '沸'):
                        point = b[i + 1]
                        break
                    else:
                        pass
            com_point=point
            ws['A'+str(k)].value=cas
            ws['B'+str(k)].value=point
            wbk.save('cas_沸点1.xlsx')
            tag=0
            print('第{0}条记录成功'.format(k))
            # sleep(1)
        except:
            print('第{0}条记录失败'.format(k),init_url)
            tag=1
            sleep(8)


# https://www.chemicalbook.com/ProductChemicalPropertiesCB8229876.htm
# https://www.chemicalbook.com/ProductChemicalPropertiesCB58227907.htm

# 沸点
# print(point)
