import json
import random
import re
import time
import urllib
import urllib.request

import openpyxl
import requests
import xlwt
import openpyxl
from openpyxl import *

# from bs4 import BeautifulSoup
# from lxml import etree

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36',

    # 'Cookie': 'navCtgScroll=0; _lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ua=dpuser_15059620632; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; fspop=test; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1634227745,1635433785,1635434475; cy=344; cye=changsha; thirdtoken=41702c09-7086-430b-ab15-e7448c57e6c0; dper=234cb94f92704aef13a7e6c816ada0ec5d82dfce20785f8f7461236e5ad4452e080a2eb9d7089f0779c7dccdb93f453054a0e3dd272f57e46cf3f29951cb4adf5159c53995424956e2bae8b3aa6d15febd25ed5e382eabe35e43e4007f6a7b94; ll=7fd06e815b796be3df069dec7836c3df; ctu=b58a9fce079b27059480a633b27a83c81b97ef52648306836b7b2a7e61a47c0127548a04b35974cf992382a35b6cbb17; dplet=7a0ee364bff4bc2bf9ed9183621d7340; _lxsdk_s=17cc773261d-562-249-600%7C%7C710; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1635436153',
    # 'Cookie': '_lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; fspop=test; ua=sunshine; dper=234cb94f92704aef13a7e6c816ada0ec99322848b9a103cfe3bd94b17b5973abe3504b1c18ec4575f82a7f8e19bc2765a23a9a9541ab953eeca862f3c2712be82cce25199a025ec99a3cb8500c39eae6f33a8983370d8fb82c9739611bd81246; ll=7fd06e815b796be3df069dec7836c3df; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1635434475,1635479818,1635574037,1635611895; cye=shanghai; cy=1; dplet=d34cef477d103bc73ac23b9303221b73; _lxsdk_s=17cd192f0f0-4b9-470-d78%7C%7C740; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1635614015',
    'Cookie': '_lxsdk_cuid=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _lxsdk=17c7f907122c8-005a3d1714b78c-b7a1438-144000-17c7f907122c8; _hc.v=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745; s_ViewType=10; ctu=1dc952f673738fa47d20413160405840235920484dc849760c4a66b318dd0f6b; cye=changsha; aburl=1; Hm_lvt_dbeeb675516927da776beeb1d9802bd4=1635694823; Hm_lvt_4c4fc10949f0d691f3a2cc4ca5065397=1635694848; fspop=test; cy=344; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1636117016,1636343453,1636353946,1636374152; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; dplet=1ef1cdd938cc644083b8f78c9dbde5d8; dper=55d202634e43c11ec7db27244d43482565ad12e25b6d1631658ad91fca2003902ca62859f247ba2013b503e0d1ade55874348cb07b6ff22f0aa1d436709f9c934c459dfe417aa3e575f97ed5fb718ee5057ce4f3ac22d30f20ec743c6f9bd46b; ll=7fd06e815b796be3df069dec7836c3df; ua=sunshine; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1636377586; _lxsdk_s=17cff35449f-e1a-f4f-f91%7C%7C1949',
    'Host': 'www.dianping.com',
    'X-Requested-With': 'XMLHttpRequest',
    # 'Referer': 'http://www.dianping.com/shop/链接'
}

# k = 1

wb=Workbook()
ws=wb.active
ws.title='银盆岭'
ws['A1']='店名'
ws['B1']='星级'
ws['C1']='人均'
ws['D1']='口味'
ws['E1']='环境'
ws['F1']='服务'
ws['G1']='网址查询'


# workbook = xlwt.Workbook()
# # 创建一个worksheet
# # sheet_name=clib_name
# worksheet = workbook.add_sheet('中山亭_乐和城')
# worksheet.write(0, 0, label='店名')
# worksheet.write(0, 1, label='星级')
# worksheet.write(0, 2, label='人均')
# worksheet.write(0, 3, label='口味')
# worksheet.write(0, 4, label='环境')
# worksheet.write(0, 5, label='服务')
# worksheet.write(0, 6, label='网址查询')

def read_excel_file():
    # r代表行数
    # 获取excle文件里面的内容
    wb = openpyxl.load_workbook('芙蓉南路沿线_地址.xlsx') #下次改成xlsx
    # print(type(wb))
    sheet = wb['芙蓉南路沿线']
    return sheet

def read_excel_data(sheet, r,c):
    # 用来查看cas_data
    # sheet.cell(row=1,column=1).value
    # **采集到要查询的数据 是str类型
    cell = sheet.cell(row=r, column=c).value
    # print(cell)
    if cell is not None:
        # 方便后期维护
        new_cell = cell
        # query_list = [new_cell]
        r += 1
        return new_cell
    else:
        return None


def find_excel_rows():
    # 用来查看表格的最大行数
    wb = openpyxl.load_workbook('芙蓉南路沿线_地址.xlsx')
    sheet = wb['芙蓉南路沿线']
    # 获取到行数
    rows = sheet.max_row
    return rows
store_content=read_excel_file()
rows=find_excel_rows()
for i in range(148,rows-1) :
    print('开始')
    try:
        store_name=read_excel_data(store_content,i+1,1)
        store_url=read_excel_data(store_content,i+1,2)
        find_url=read_excel_data(store_content,i+1,3)
        # url = 'https://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=l6y9rHNsA0KdKUMi&cityId=344&mainCategoryId=34245'
        # time.sleep(random.randint(1, 2))
        # request = urllib.request.Request(url=url, headers=headers)
        # response = opener.open(request)
        # response = urllib.request.urlopen(request)
        response = requests.get(url=store_url, headers=headers)
        print(response)
        content = response.text
        print(find_url)
        # print(content)
        dic = {'&#xefb0': '0', '&#xe7c6': '1', '&#xe9e5': '2', '&#xe838': '3', '&#xe364': '4', '&#xf128': '5',
               '&#xf639': '6', '&#xf440': '7', '&#xe174': '8', '&#xe7c3': '9', }
        # items=soup.find_all(class_='brief-info')

        obj = json.loads(content)
        print('obj加载成功！')
        print(obj)
        lenth = len(obj["avgPrice"].split('>'))
        obj_price = obj["avgPrice"]
        if(obj_price[0]=='0'):
            price='--(没有)'
        else:
            # 处理价格
            if (lenth == 3):
                b = obj_price.split('>')[1].split(';')[0]
                b = dic[b]
                # 要加‘1’  注意这个没有处理单价为1元
                if (obj_price[0] == '1'):
                    if (obj_price[-1] == '1'):  # 处理101的情况
                        price = '1' + b + '1' + '元'
                    # 处理110
                    elif (obj_price[1] == '1'):
                        price = '11' + b + '元'
                    else:
                        price = '1' + b + '元'
                elif (obj_price[-1] == '1'):  # 处理011
                    if (obj_price[-2] == '1'):
                        price = b + '11' + '元'
                    else:
                        price = b + '1' + '元'
                else:  # 处理个位数情况
                    price = b + '元'

            elif (lenth == 5):
                # 处理221  212 22 122
                # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xf5d6;</d', '']
                a = obj_price.split('>')[1].split(';')[0]
                a = dic[a]
                b = obj_price.split('>')[3].split(';')[0]
                b = dic[b]
                # print(obj_price.split('>'))
                # print(obj_price.split('>')[0][0])
                # 122
                if (obj_price.split('>')[0][0] == '1'):
                    price = '1' + a + b + '元'
                    # print('11111')
                # 221
                elif (obj_price.split('>')[2][0] == '1'):
                    price = a + b + '1' + '元'
                    # print('2222222')
                # 212
                elif (obj_price.split('>')[4] == '1'):
                    price = a + '1' + b + '元'
                    # print('3333333')
                else:
                    # print('44444')
                    price = a + b + '元'
            elif (lenth == 7):
                # 处理000的情况
                # o='<d class=\"num\">&#xf5d6;</d><d class=\"num\">&#xeedb;</d><d class=\"num\">&#xf26a;</d>'
                # ['<d class="num"', '&#xf5d6;</d', '<d class="num"', '&#xeedb;</d', '<d class="num"', '&#xf26a;</d', '']
                a = obj_price.split('>')[1].split(';')[0]
                a = dic[a]
                b = obj_price.split('>')[3].split(';')[0]
                b = dic[b]
                c = obj_price.split('>')[5].split(';')[0]
                c = dic[c]
                price = a + b + c + '元'
            else:
                price = '--(没有)'
        print(price)
        # 评分
        if('fiveScore' not in obj.keys()):
            obj_sore='--(没有)'
        else:
            obj_sore = obj["fiveScore"]
        if(obj["shopRefinedScoreValueList"][0]=='-'):
            obj_taste='--(没有)'
        else:
            # 口味
            obj_taste_1 = obj["shopRefinedScoreValueList"][0].split('>')[1].split(';')[0]
            obj_taste_1 = dic[obj_taste_1]
            # obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
            obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')
            if (len(obj_taste_2) == 3):
                obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[2]
            else:
                obj_taste_2 = obj["shopRefinedScoreValueList"][0].split('>')[3].split(';')[0]
            obj_taste_2 = dic[obj_taste_2]
            obj_taste = obj_taste_1 + '.' + obj_taste_2
        if(obj["shopRefinedScoreValueList"][1]=='-'):
            obj_env='--(没有)'
        else:
            # 环境
            obj_env_1 = obj["shopRefinedScoreValueList"][1].split('>')[1].split(';')[0]
            obj_env_1 = dic[obj_env_1]
            # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
            # obj_env_2=obj["shopRefinedScoreValueList"][1].split('>')[2]
            obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')
            if (len(obj_env_2) == 3):
                obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[2]
            else:
                obj_env_2 = obj["shopRefinedScoreValueList"][1].split('>')[3].split(';')[0]
            # print(obj_env_2)
            obj_env_2 = dic[obj_env_2]
            obj_env = obj_env_1 + '.' + obj_env_2
        if(obj["shopRefinedScoreValueList"][2]=='-'):
            obj_serve='--(没有)'
        else:
            # 服务
            obj_serve_1 = obj["shopRefinedScoreValueList"][2].split('>')[1].split(';')[0]
            obj_serve_1 = dic[obj_serve_1]
            # obj_serve_2=obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
            obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')
            if (len(obj_serve_2) == 3):
                obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[2]
            else:
                obj_serve_2 = obj["shopRefinedScoreValueList"][2].split('>')[3].split(';')[0]
            # print(obj_serve_2)
            obj_serve_2 = dic[obj_serve_2]
            # print(obj_serve_2)
            obj_serve = obj_serve_1 + '.' + obj_serve_2
        print(obj_sore, price, obj_taste, obj_env, obj_serve)


        ws['A'+str(i+2)].value=store_name
        ws['B'+str(i+2)].value=obj_sore
        ws['C'+str(i+2)].value=price
        ws['D'+str(i+2)].value=obj_taste
        ws['E'+str(i+2)].value=obj_env
        ws['F'+str(i+2)].value=obj_serve
        ws['G'+str(i+2)].value=find_url
        # worksheet.write(i+1, 0, label=store_name)
        # worksheet.write(i+1, 1, label=obj_sore)
        # worksheet.write(i+1, 2, label=price)
        # worksheet.write(i+1, 3, label=obj_taste)
        # worksheet.write(i+1, 4, label=obj_env)
        # worksheet.write(i+1, 5, label=obj_serve)
        # worksheet.write(i+1, 6, label=find_url)

        # request=urllib.request.Request(url=src,headers=headers)
        # content=get_content(request)
        # tree=etree.HTML(content)
        # //div[@class="mid-score score-45"]
        # mark_list=tree.xpath('//div[@class="star-wrapper"]/div')
        # 评分地址http://www.dianping.com/ajax/json/shopDynamic/reviewAndStar?shopId=H9Vnbc8KMYG56vZN&cityId=344&mainCategoryId=34246&_token=eJx1kM1OwkAUhd9l1pN2%2FilNXEA0UKBqSqkUw6IdsCXQUpgKqPHdvaO4cOHqfHPm3NyT%2B4GOwQr5lBAiKEan9RH5iDrEUQij1sCP4lJwj1DGGMdI%2F%2FWUHcqPyS3yn6nkCnMql9aJwPhxPCWW%2BIoMkAlsBeUBRFDZto3vuufz2VltsrrZ1IWj95Vryn3jDrtJnWtvHKYDqU6Le6j0X16XWV2YMgOgxC0o7SDYUcV2B%2Bt2oBeFaeZ5mEnxTRJqMEsdhmmX2%2FjWxkGzq7a%2F7xCOAlGzKWqg9egST40wh5coNPGMhclwEuv3UBNhpjP%2BOBFBMFhU6SFKatnXvYdIPaXz3egtytenZvjaVIvdbC7HvT5jd9v0Bn1%2BAc%2BGbCo%3D&uuid=751dd7ad-f61c-e41b-0401-95d1866ae493.1634227745&platform=1&partner=150&optimusCode=10&originUrl=http%3A%2F%2Fwww.dianping.com%2Fshop%2FH9Vnbc8KMYG56vZN
        print('第{0}条数据成功记录'.format(i))
        wb.save('芙蓉南路沿线数据3.xlsx')
        print()
    except:
        wb.save('芙蓉南路沿线数据3.xlsx')
        print(store_url)
        print('第{0}条数据处理失败'.format(i))




# "1<d class=\"num\">&#xe405;</d>"  1开头
# http://www.dianping.com/shop/H8qpxHT6tl72i2ME 单独处理 107


# else:
#     pass
# http://www.dianping.com/shop/G6cjEzJiVFGuY8Vm
# http://www.dianping.com/shop/FujE6yfc4atZpm5K
# if __name__ == '__main__':

# workbook.save('长沙test.xls')
