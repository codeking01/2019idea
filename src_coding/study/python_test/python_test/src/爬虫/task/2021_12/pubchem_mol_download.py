'''
@author king_xiong
@date 2021-12-16 20:02
'''
import os
import time
import urllib.request
import openpyxl
import requests
import json
import pymysql
#页面地址
# https://pubchem.ncbi.nlm.nih.gov/compound/5288826
# 下载地址
# https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/5288826/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_5288826
headers={
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.20 Safari/537.36',
}
proxies={
    # 这个位置加代理ip
}
# import pymysql
# # 打开数据库连接
# db = pymysql.connect(host="localhost", user="root", password="123456", database="417database")
# # 使用cursor（）方法获取游标
# cursor=db.cursor()
# # 执行SQL语句
# insert_sql = "insert into test01(id, name, age) values(%s, %s, %s)"
# parm=(4,"是的",2)
# # 执行sql语句
# cursor.execute(insert_sql,parm)
# # 提交到数据库执行
# db.commit()
# cursor.execute("select * from test01")
# # 查看表里所有数据
# data = cursor.fetchall()
# print(data)
if __name__ == '__main__':
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['CAS']
    # 读取excel中的最大行数
    rows = sheet.max_row
    for t in range(rows + 1):
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=t+2, column=1).value
            deal_cas=cas.replace('_','-')
            # 下面这个是为了获取需要链接的cid号
            url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
            deal_cas) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
            response = requests.get(url=url, headers=headers, proxies=proxies)
            response.encoding = 'utf-8'
            # 下面这个json数据可以采集到
            json_content = response.text
            # print(page_content)
            # ***可以直接读取json，就不存进去了
            obj = json.loads(json_content)
            # 找到json中相应的数据
            obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
            # 获取下载链接https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/5288826/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_5288826
            url = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/{obj_cid}'.format(obj_cid=obj_cid)+'/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_{obj_cid}'.format(obj_cid=obj_cid)
            urllib.request.urlretrieve(url, './mol_files_pubchem/' + str(cas) + '.sdf')
            print('第{0}条数据处理成功'.format(t))
            time.sleep(10)
            if(t%500==0 and t!=0):
                print('休息500s,不要着急，小心封号！')
                time.sleep(500)
        except:
            # 失败播放音乐
            file_path = "C:\\Users\\king\Downloads\Music\错位时空.mp3"
            os.startfile(file_path)
            time.sleep(10)
            print('*********第{0}条数据处理失败....'.format(t))
            print('---------失败的cas为'+str(deal_cas))