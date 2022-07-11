'''
@author king_xiong
@date 2021-12-16 21:20
'''
# 下载链接
# https://webbook.nist.gov/cgi/cbook.cgi?Str2File=C100005
import os
import time
import urllib.request
import openpyxl
import requests

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.20 Safari/537.36',
}
proxies = {
    # 这个位置加代理ip
}
# import pymysql
# # 打开数据库连接
# db = pymysql.connect(host="localhost", user="root", password="123456", database="417database")
# # 使用cursor（）方法获取游标
# cursor=db.cursor()
# # 执行SQL语句
# insert_sql = "insert into test01(id, name, age) values(%s, %s, %s)"
# parm=(4,"是的",2)
# # 执行sql语句
# cursor.execute(insert_sql,parm)
# # 提交到数据库执行
# db.commit()
# cursor.execute("select * from test01")
# # 查看表里所有数据
# data = cursor.fetchall()
# print(data)

if __name__ == '__main__':
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['CAS']
    # 读取excel中的最大行数
    rows = sheet.max_row
    for t in range(14264, rows + 1):
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=t + 2, column=1).value
            # 获取操作的Str2File（需要去掉_）
            deal_cas = cas.replace('_', '')
            url = 'https://webbook.nist.gov/cgi/cbook.cgi?Str2File=C{0}'.format(deal_cas)
            urllib.request.urlretrieve(url, './mol_files_NIST/' + str(cas) + '.mol')
            print('第{0}条数据处理成功'.format(t))
            # time.sleep(10)
            if (t % 1000 == 0 and t != 0):
                time.sleep(60)
        except:
            # 失败播放音乐
            file_path = "C:\\Users\\king\Downloads\Music\错位时空.mp3"
            os.startfile(file_path)
            time.sleep(10)
            print('*********第{0}条数据处理失败'.format(t))
            print('---------失败的cas为' + str(cas))
    print('爬取完成了！！')
